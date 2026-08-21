import json
from openpyxl import load_workbook

# Step A: Excel file open karo, read-only mode mein (fast, memory-safe)
wb = load_workbook("data/MoinSystems_AI_Public_Chatbot_RAG_Dataset_v2.xlsx", read_only=True)
ws = wb["RAG_Knowledge"]   # jo sheet mein actual knowledge data hai

# Step B: Saari rows ek list mein le lo
rows = list(ws.iter_rows(values_only=True))
header = rows[0]           # pehli row = column names
data_rows = rows[1:]       # baaki sab actual records

# Step C: Har row ko dictionary + JSON line mein convert karo
output = []
for row in data_rows:
    rid, title, category, tags, intents, content, embed_text, status, source_basis = row[:9]

    if not rid:                     # khaali rows skip
        continue
    if status != "cleaned_validated":  # sirf APPROVED data lo, unverified skip
        continue

    record = {
        "id": rid,
        "title": title,
        "category": category,
        "tags": [t.strip() for t in (tags or "").split(",") if t.strip()],
        "intents": [i.strip() for i in (intents or "").split(",") if i.strip()],
        "content": content,
        "text": embed_text,   # ye field embed hogi
        "metadata": {
            "dataset_version": "v2",
            "data_status": status,
            "source_basis": source_basis
        }
    }
    output.append(record)

# Step D: JSONL file likho — ek record, ek line
with open("data/MoinSystems_AI_Public_Chatbot_RAG_Dataset_v2.jsonl", "w") as f:
    for rec in output:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")

print(f"Wrote {len(output)} records to JSONL")